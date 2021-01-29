import requests
import pprint
import jsonpath
import openpyxl
# -----------------------------------------------------------------------------------------------------------
# def register_auto(url,register_data):
#     url_address = url
#     request_data = register_data
#     head = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
#     response = requests.post(url = url_address,json = request_data,headers= head)
#     return response.json()
# pprint.pprint(ask)
# url = "http://8.129.91.152:8766/futureloan/member/register"
# register_data = {
#   "mobile_phone": "15815541763",
#   "pwd": "lemon123456",
#   "type":"0",
#   "reg_name":"管理员用户lemon"
# }
# final_result = register_auto(url,register_data)
# # member_id = final_result["data"]["id"] or jsonpath.jsonpath(final_result,"$.data.id")[0]
# # token_id = final_result["data"]["token_info"]["token"] or jsonpath.jsonpath(final_result,"$.data.token_info.token)[0]
# pprint.pprint(final_result)
# -------------------------------------------------------------------------------------------------------------------
# def data_abstract(filename,sheetname):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb[sheetname]
#     max_row = sheet.max_row# 取到尽头
#     cases =[]
#     for i in range(2,max_row+1):
#         filecase = dict(
#         case_id=sheet.cell(row=i, column=1).value,
#         url = sheet.cell(row=i, column=5).value,
#         data =sheet.cell(row=i, column=6).value,
#         expected_result =sheet.cell(row=i, column=7).value)
#         cases.append(filecase)
#     return cases
# substitution = data_abstract("test_case_api.xlsx","register")
# pprint.pprint(substitution)
# -------------------------------------------------------------------------------------------------------------------
# def data_writting(filename,sheetname,revised,row,column):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb[sheetname]
#     sheet.cell(row,column).value = revised
#     wb.save(filename)
#     return  revised
# writting = data_writting("test_case_api.xlsx","register",revised,row,column)
# pprint.pprint(writting)
#重新做 -------------------------------------------------------------------------------------------------------------
def function_display(url,function_data):
    url1 = url
    data1 = function_data
    head1 = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    response1 = requests.post(url = url1,json = data1,headers =head1)
    return response1.json()
# -------------------------------------------------------------------------------------------------------------------
def data_abstract1(filename,sheetname):
    wb1 = openpyxl.load_workbook(filename)
    sheet1 =wb1[sheetname]
    max_row1 = sheet1.max_row
    # print(max_row1)
    cases1 =[]
    for i in range(2,max_row1+1):
        dict1 = dict(
        case_id1 = sheet1.cell(row =i,column=1).value,
        sheet1_url = sheet1.cell(row =i,column=5).value,
        sheet1_data = sheet1.cell(row =i,column=6).value,
        sheet1_expection = sheet1.cell(row =i,column=7).value)
        # print(sheet1_expection)
        cases1.append(dict1)
    return cases1
# data_abstract1("test_case_api.xlsx","register")
    # print(cases1)
# -------------------------------------------------------------------------------------------------------------------
def data_writting1(filename,sheetname,row,column,data_return):
    wb1 = openpyxl.load_workbook(filename)
    sheet1 =wb1[sheetname]
    sheet1.cell(row,column).value =data_return
    wb1.save(filename)
    return data_return

if __name__ == '__main__':
    result = data_abstract1("test_case_api.xlsx","register")
    print(result)
# -------------------------------------------------------------------------------------------------------------------

