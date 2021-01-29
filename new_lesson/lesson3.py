#自动化实现结果对比并且写入Excel
#自动读取Excel里面的测试用例 用openpyxl
#需要的读取和写入的Excel文档加载进来----load workbook====放到同一级目录
#再获取表格对象
#表单对象
import requests
import openpyxl

#     wb = openpyxl.load_workbook("test_case_api.xlsx")  # 获取所需要的表单
#     sheet = wb["register"]
#     cell = sheet.cell(row=1, column=1)  # 找到表中所需数据单元格位置
# for i in range(2,15):
#     url_api = sheet.cell(row=i,column=5)
#     api_data = sheet.cell(row=i,column=5)
#     expected_result = sheet.cell(row=i,column=6)
# import pprint #更加明显显示如jmeter一样的内容
# -------------------------------------------------------------------------------------------------------------
# 把接口请求设置为一个函数
def api_fun(url_api,api_data):
    head = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    repsons = requests.post(url=url_api,json=api_data,headers=head)
    return repsons.json()

# print(repsons.json())
# res = repsons.json()#这里面的就是我们要的包含结果的数据
# pprint.pprint(repsons.json())


url_api = "http://8.129.91.152:8766/futureloan/member/register"
api_data = {
      "mobile_phone":"15815541767",
      "pwd":"lemon123456",
      "type":"1",
      "reg_name":"lllllemon"
    }
resp = api_fun(url_api,api_data)
print(resp)
# --------------------------------------------------------------------------------------------------------------------
# 把获取的excel列表的公式改为函数调取
import requests
import openpyxl
def data_reading(filename,sheetname):
    wb = openpyxl.load_workbook(filename)  # 获取所需要的表单
    sheet = wb[sheetname]
    cases = []
    max_row = sheet.max_row
    for i in range(2,max_row+1):
        case = dict(
        url_api= sheet.cell(row=i,column=5).value,
        api_data= sheet.cell(row=i,column=5).value,
        expected_result = sheet.cell(row=i,column=6).value)
        cases.append(case)
    return cases
result = data_reading("test_case_api.xlsx","register")
# --------------------------------------------------------------------------------------------------------------------
# 实际结果写入
def data_write(filename,sheetname,revised,rowx):
    wb = openpyxl.load_workbook(filename)  # 获取所需要的表单
    sheet = wb[sheetname]
    sheet.cell(rowx, column=8).value = revised  # 找到表中所需数据单元格位置
    wb.save("test_case_api.xlsx")
    return revised




